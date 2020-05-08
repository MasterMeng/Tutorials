import random
import datetime
import os
from openpyxl import Workbook, load_workbook

# 全局变量
# 定义列名
titles = ['序号', '部门', '姓名', '目前所在地', '是否返回', '返程方式', '返程时间']
ways = {0: '自驾', 1: '高铁', 2: '飞机'}

'''
@param filename:生成的excel文件名
'''
def CreateExcel(filename):
    # 创建文件对象
    wb = Workbook()

    # 获取一个sheet
    firstSheet = wb.active

    # 写入内容
    # 这里选择使用append()方法，该方法直接将列表作为一行添加到execl中，列表中的每个元素存入一个一个单元格
    firstSheet.append(titles)

    # 按单元格写入excel
    # 由于上面已经写入了一行内容，所以这里从第2行开始
    firstSheet['A2'] = 1
    '''
    random.randint(1, 10) 返回1到10之间的整数
    '''
    firstSheet['B2'] = '部门'+str(random.randint(1, 10))
    firstSheet['C2'] = 'name'+str(random.randint(1, 10))
    firstSheet['D2'] = '地址'+str(random.randint(1, 10))
    if random.randint(0, 3000) % 2 == 0:
        firstSheet['E2'] = '是'
    else:
        firstSheet['E2'] = '否'
    firstSheet['F2'] = ways[random.randint(1, 9999) % 3]
    '''
    datetime.date.today() 获取当前日期，
    strftime('%Y-%m-%d') 格式化为“年-月-日”格式的字符串
    '''
    firstSheet['G2'] = datetime.date.today().strftime('%Y-%m-%d')

    # 保存
    wb.save(filename)


'''
@param files: 要汇总的excel文件列表
@param filename: 汇总后的excel文件名
'''
def CountExcel(files, filename):
    # 打开文件对象
    wb = Workbook()
    # 获取一个sheet
    ws = wb.active
    # 写入表头
    ws.append(titles)

    i = 1
    for name in files:
        # 打开已存在的文件
        wbTmp = load_workbook(name)
        wsTmp = wbTmp.active
        rows = []
        # 迭代获取sheet中的所有行
        for row in wsTmp.iter_rows():
            rows.append(row)

        # 为序号赋值
        rows[1][0].value = i

        # 因为每个excel表都只有两行，且第一行为表头，所以这里直接取第2行中的单元格的值
        cells = []
        for r in rows[1]:
            cells.append(r.value)
        ws.append(cells)
        i += 1

    wb.save(filename)


if __name__ == "__main__":
    # 删除当前目录下的total.xlsx
    if os.path.exists('total.xlsx'):
        os.remove('total.xlsx')

    # 生成200个excel文件
    for i in range(1, 201):
        CreateExcel(str(i)+'.xlsx')

    excels = []
    # 遍历当前文件夹，获取目录下的所有excel文件
    for filepath, dirname, filenames in os.walk(r'.'):
        # 根据后缀来判断excel
        for filename in filenames:
            if os.path.splitext(filename)[-1] == '.xlsx':
                excels.append(filename)

    # 将所有的excel文件整合成一个文件
    CountExcel(excels, 'total.xlsx')

    # for excel in excels:
    #     os.remove(excel)
